import csv, io

import openpyxl
from django.db import transaction
from django.db.models import Sum
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.contrib.auth.hashers import make_password

# Create your views here.
# one parameter named request
from ..models import *
# from django.contrib.auth.models import User
from django.contrib.auth import get_user_model

User = get_user_model()


def subject_result_upload(request, get__subject, rank):
    get_rank = get_object_or_404(Rank, name=rank)
    check_subject = get_object_or_404(Subject, name=get__subject)
    if request.method == "GET":
        return redirect('SRS:subject_results_list', rank=get_rank, subject_name=check_subject)
    else:
        excel_file = request.FILES['excel_file']

        # you may put validations here to check extension or file size

        wb = openpyxl.load_workbook(excel_file)

        # getting a particular sheet by name out of many sheets
        worksheet = wb.worksheets[0]
        get_event = AcademicEvent.objects.get(rank=get_rank, is_active=True)

        # print(worksheet)

        # iterating over the rows and
        # getting value from each cell in row
        # try:

        for rowno, rowval in enumerate(worksheet.iter_rows(min_row=2, max_row=worksheet.max_row), start=2):
            # if check_subject.name == column[4] and get_rank.name == column[2]:

            registration_number = Registration.objects.filter(
                student__admission=worksheet.cell(row=rowno, column=2).value).first()
            check_combination = CombinationSubject.objects.get(subject=check_subject,
                                                               combination=registration_number.combination)
            with transaction.atomic():
                card = Result.objects.select_for_update().filter(registration=registration_number,
                                                                  event=get_event,
                                                                  subject=check_combination,
                                                                  )
                if card:
                    for created in card:
                        created.marks=int(worksheet.cell(row=rowno, column=7).value)
                        created.save()

                        created.save()
                        if created:
                            x = 7
                            y = 3

                            get_total_subject = Result.objects.filter(registration=registration_number, event=get_event).count()
                            get_subjects = registration_number.combination.subject
                            if get_total_subject == get_subjects and get_subjects >= 7:

                                with transaction.atomic():
                                    cards = Result.objects.select_for_update().filter(registration=registration_number,
                                                                                      event=get_event, marks__gte=0
                                                                                      ).order_by('-marks')[:x]
                                    if cards:
                                        total = 0
                                        # total_weight = 0

                                        for mark in cards:
                                            total = total + mark.point
                                            # total_weight = total_weight + mark.marks

                                        save_data = YearResult(
                                            registration=registration_number,
                                            event=get_event,
                                            point=total
                                            # weight = total_weight
                                        )
                                        save_data.save()
                            elif get_total_subject == get_subjects and get_subjects <= 5:

                                with transaction.atomic():
                                    cards = Result.objects.select_for_update().filter(registration=registration_number,
                                                                                      event=get_event, marks__gte=0,
                                                                                      subject__subject__is_core=True
                                                                                      ).order_by('-marks')[:y]
                                    if cards:
                                        total = 0
                                        # total_weight = 0

                                        for mark in cards:
                                            total = total + mark.point
                                            # total_weight = total_weight + mark.marks

                                        save_data = YearResult(
                                            registration=registration_number,
                                            event=get_event,
                                            point=total
                                            # weight = total_weight
                                        )
                                        save_data.save()
        messages.success(request,
                         f"Result Uploaded Successfully")
        return redirect('SRS:subject_results_list', rank=get_rank, subject_name=check_subject)


def student_assessment_result_upload(request, rank, combination):
    get_rank = get_object_or_404(Rank, name=rank)
    get_combination = get_object_or_404(Combination, name=combination)
    if request.method == "GET":
        return redirect('SRS:student_character', rank=get_rank, combinatione=get_combination)
    else:
        excel_file = request.FILES['excel_file']

        # you may put validations here to check extension or file size

        wb = openpyxl.load_workbook(excel_file)

        # getting a particular sheet by name out of many sheets
        worksheet = wb.worksheets[0]
        # get_event = AcademicEvent.objects.get(rank=get_rank, is_active=True)

        # print(worksheet)

        # iterating over the rows and
        # getting value from each cell in row
        # try:

        for rowno, rowval in enumerate(worksheet.iter_rows(min_row=2, max_row=worksheet.max_row), start=2):
            # if check_subject.name == column[4] and get_rank.name == column[2]:

            registration_number = Registration.objects.filter(
                student__admission=worksheet.cell(row=rowno, column=2).value).first()

            create_901 = StudentCharacter.objects.create(
                registration=registration_number,
                character=Character.objects.get(code="901"),

                grade=worksheet.cell(row=rowno, column=6).value

            )

            create_902 = StudentCharacter.objects.create(
                registration=registration_number,
                character=Character.objects.get(code="902"),

                grade=worksheet.cell(row=rowno, column=7).value

            )

            create_903 = StudentCharacter.objects.create(
                registration=registration_number,
                character=Character.objects.get(code="903"),

                grade=worksheet.cell(row=rowno, column=8).value
            )
            create_904 = StudentCharacter.objects.create(
                registration=registration_number,
                character=Character.objects.get(code="904"),

                grade=worksheet.cell(row=rowno, column=9).value

            )
            create_905 = StudentCharacter.objects.create(
                registration=registration_number,
                character=Character.objects.get(code="905"),

                grade=worksheet.cell(row=rowno, column=10).value
            )
            create_906 = StudentCharacter.objects.create(
                registration=registration_number,
                character=Character.objects.get(code="906"),

                grade=worksheet.cell(row=rowno, column=11).value

            )
            create_907 = StudentCharacter.objects.create(
                registration=registration_number,
                character=Character.objects.get(code="907"),

                grade=worksheet.cell(row=rowno, column=12).value

            )
            create_908 = StudentCharacter.objects.create(
                registration=registration_number,
                character=Character.objects.get(code="908"),

                grade=worksheet.cell(row=rowno, column=13).value

            )
            create_909 = StudentCharacter.objects.create(
                registration=registration_number,
                character=Character.objects.get(code="909"),

                grade=worksheet.cell(row=rowno, column=14).value

            )
            create_910 = StudentCharacter.objects.create(
                registration=registration_number,
                character=Character.objects.get(code="910"),

                grade=worksheet.cell(row=rowno, column=15).value

            )
            create_911 = StudentCharacter.objects.create(
                registration=registration_number,
                character=Character.objects.get(code="911"),

                grade=worksheet.cell(row=rowno, column=16).value

            )

        messages.success(request,
                         f"Result Uploaded Successfully")
        return redirect('SRS:student_character', rank=get_rank, combination=get_combination)


def import_staff(request):
    if request.method == "GET":
        return redirect('SRS:staff_entry')
    else:
        excel_file = request.FILES['excel_file']

        # you may put validations here to check extension or file size

        wb = openpyxl.load_workbook(excel_file)

        # getting a particular sheet by name out of many sheets
        worksheet = wb.worksheets[0]

        # print(worksheet)

        # iterating over the rows and
        # getting value from each cell in row
        # try:

        for rowno, rowval in enumerate(worksheet.iter_rows(min_row=2, max_row=worksheet.max_row), start=2):
            # if check_subject.name == column[4] and get_rank.name == column[2]:

            User.objects.create(
                email=worksheet.cell(row=rowno, column=1).value,
                first_name=worksheet.cell(row=rowno, column=2).value,
                middle_name=worksheet.cell(row=rowno, column=3).value,
                last_name=worksheet.cell(row=rowno, column=4).value,
                sex=worksheet.cell(row=rowno, column=5).value,
                phone=worksheet.cell(row=rowno, column=6).value,
                title=worksheet.cell(row=rowno, column=7).value,
                password=make_password(worksheet.cell(row=rowno, column=4).value),
                is_staff=True

            )

        messages.success(request,
                         f"Staff Imported Successfully")
        return redirect('SRS:staff_entry')
